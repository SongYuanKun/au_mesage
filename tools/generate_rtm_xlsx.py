from __future__ import annotations

import pathlib
from typing import Any

import openpyxl
import yaml


ROOT = pathlib.Path(__file__).resolve().parents[1]
REQ_PATH = ROOT / "docs" / "SRS" / "requirements.yml"
REFS_PATH = ROOT / "docs" / "SRS" / "rtm_refs.yml"
OUT_DIR = ROOT / "docs" / "SRS" / "attachments" / "rtm"


def _as_list(v: Any) -> list:
    if v is None:
        return []
    if isinstance(v, list):
        return v
    return [v]


def _flatten_sources(req: dict[str, Any]) -> str:
    sources = _as_list(req.get("sources"))
    return "\n".join(str(s) for s in sources) if sources else ""


def _flatten_risks(req: dict[str, Any]) -> str:
    risks = _as_list(req.get("risks"))
    return "\n".join(str(r) for r in risks) if risks else ""


def _load_refs() -> dict[str, Any]:
    if not REFS_PATH.is_file():
        return {}
    return yaml.safe_load(REFS_PATH.read_text(encoding="utf-8")) or {}


def _ref_fields(req_id: str, refs: dict[str, Any]) -> tuple[str, str, str]:
    entry = refs.get(req_id) or {}
    design = entry.get("design") or "TBD"
    code = entry.get("code") or "TBD"
    test = entry.get("test") or "TBD"
    return str(design), str(code), str(test)


def _write_sheet(ws, items: list[dict[str, Any]], req_type: str, refs: dict[str, Any]) -> None:
    headers = [
        "ReqID",
        "Name",
        "Type",
        "Module/Category",
        "MoSCoW",
        "Effort",
        "SourceRefs",
        "DesignRef",
        "CodeRef",
        "TestRef",
        "Owner",
        "Status",
        "Risks",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="EDEDED")

    for r in items:
        module = r.get("module") or r.get("category") or r.get("type") or "-"
        req_id = str(r.get("id", "-"))
        design_ref, code_ref, test_ref = _ref_fields(req_id, refs)
        status = "InProgress" if "TBD" not in (design_ref, code_ref, test_ref) else "Draft"
        ws.append(
            [
                req_id,
                r.get("name", "-"),
                req_type,
                module,
                r.get("moscow", "-"),
                r.get("effort", "-"),
                _flatten_sources(r),
                design_ref,
                code_ref,
                test_ref,
                "TBD",
                status,
                _flatten_risks(r),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"
    widths = {
        "A": 14,
        "B": 30,
        "C": 10,
        "D": 22,
        "E": 10,
        "F": 10,
        "G": 40,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 12,
        "L": 10,
        "M": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")


def main() -> None:
    data = yaml.safe_load(REQ_PATH.read_text(encoding="utf-8"))
    refs = _load_refs()

    fr = data.get("functional_requirements", [])
    nfr = data.get("non_functional_requirements", [])
    interfaces = data.get("interfaces", [])
    sec = data.get("security_compliance", [])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_fr = wb.create_sheet("FR")
    _write_sheet(ws_fr, fr, "FR", refs)

    ws_nfr = wb.create_sheet("NFR")
    _write_sheet(ws_nfr, nfr, "NFR", refs)

    ws_if = wb.create_sheet("Interfaces")
    _write_sheet(ws_if, interfaces, "IF", refs)

    ws_sec = wb.create_sheet("Security")
    _write_sheet(ws_sec, sec, "SEC", refs)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "RTM.xlsx"
    wb.save(out_path)


if __name__ == "__main__":
    main()

